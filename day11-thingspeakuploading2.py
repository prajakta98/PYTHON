import requests
from openpyxl import Workbook
import matplotlib
import matplotlib.pyplot as plt
from pylab import *
r=requests.get("https://api.thingspeak.com/channels/527215/feeds.json?api_key=CHTAARP7OEKCRSVY&results=10")
print(r.status_code)
#print(r.json())
channelDict=r.json()
#print(channelDict)
feedsList=channelDict['feeds']
print(feedsList[0])
field1Values=[]
for everyitem in feedsList:
    field1Values.append(int(everyitem['field1']))
print(field1Values)
wb=Workbook()
ws1=wb.active
ws1.title="ThingSpeak Data"
ws1['A1'],ws1['B1']="Sr.No","Field1"
print(wb.sheetnames)
print(field1Values)
count=2
for everyitem in field1Values:
    ws1.cell(row=count,column=2).value=everyitem
    count+=1
count=1
for everyno in range(2,12):
    ws1.cell(row=everyno,column=1).value=count
    count+=1
wb.save("ThingSpeak Channel Databook.xlsx")
field1Values_nparray=np.array(field1Values)
print(field1Values_nparray)
xvalues=np.array(list(range(1,11)))
print(xvalues)
fig,axes=plt.subplots(figsize=(12,4))
ax=axes
ax.plot(xvalues,field1Values_nparray,color='red',marker='o')
plt.show()
fig.savefig("ThingSpeakGraph",dpi=500)