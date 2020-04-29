from openpyxl import load_workbook
from openpyxl.utils import get_column_letter,column_index_from_string
wb=load_workbook('VITPython18-FinalProject-Workbook.xlsx')
sheet=wb["UserDetails"]
print(sheet.title)
username=input("Enter Full Name:")
def attempts1():
    for j in range(0,3):
        print("invalid \n Try Again")
        if j<2:
            username=input("Enter Full Name:")
            n=checkname()
        else:
            print("program quit in stage 1")
def attempts2():
    for k in range(0,3):
        print("invalid \n Try Again")
        if k<2:
            accno=input("Enter Account Number:")
            m=checkaccno()
        else:
            print("program quit in stage 1")
def checkname(n):
    for i in range(2,7):
        if username in sheet.cell(row=i,column=4).value:
            print("valid")
            return 0
        else:
            i=i+1
            a=attempts1()
            return 1
def checkaccno(m):
    for i in range(2,7):
        if accno in sheet.cell(row=i,column=3).value:
            print("valid")
            return 0
        else:
            i=i+1
            return 1
            b=attempts2()
n=checkname()
#if n==0:
   # m=checkacc
       # return 2
    # elif username in sheet.cell(row=3,column=4).value:
        # print("valid")
        # return 3
    # elif username in sheet.cell(row=4,column=4).value:
        # print("valid")
        # return 4
    # elif username in sheet.cell(row=5,column=4).value:
        # print("valid")
        # return 5
    # elif username in sheet.cell(row=6,column=4).value:
        # print("valid")
        # return 6
    # else:
        # return 0
# def checkaccno():
    # if accountno in sheet.cell(row=2,column=3).value:
        # print("valid")
        # return 2
    # elif accountno in sheet.cell(row=3,column=3).value:
        # print("valid")
        # return 3
    # elif accountno in sheet.cell(row=4,column=3).value:
        # print("valid")
        # return 4
    # elif accountno in sheet.cell(row=5,column=3).value:
        # print("valid")
        # return 5
    # elif accountno in sheet.cell(row=6,column=3).value:
        # print("valid")
        # return 6
    # else:
	    # return 0
# n=checkname()
# if n==2 or n==3 or n==4 or n==5 or n==6:
    # accountno=input("enter a/c no:")
    # m=checkaccno()
# elif n==3:
    # sheet['G3']="Blocked"
    # print("Account is blocked")
# elif n==4:
    # sheet['G4']="Blocked"
    # print("Account is blocked")
# elif n==5:
    # sheet['G5']="Blocked"
    # print("Account is blocked")
# elif n==6:
    # sheet['G6']="Blocked"
    # print("Account is blocked")
# else:
    # for i in range(0,3):
        # print("invalid \n Try Again")
        # if i<2:
            # username=input("Enter Full Name:")
            # n=checkname()
        # else:
            # print("program quit in stage 1")
# if m==2:
    # sheet['G2']="Blocked"
    # print("Account is blocked")
# elif m==3:
    # sheet['G3']="Blocked"
    # print("Account is blocked")
# elif m==4:
    # sheet['G4']="Blocked"
    # print("Account is blocked")
# elif m==5:
    # sheet['G5']="Blocked"
    # print("Account is blocked")
# elif m==6:
    # sheet['G6']="Blocked"
    # print("Account is blocked")
# else:
    # for i in range(0,3):
        # print("invalid \n Try Again")
        # if i<2:
            # accountno=input("Enter a/c no:")
			# m=checkacccno()
        # else:
            # print("program quit in stage 1")
      