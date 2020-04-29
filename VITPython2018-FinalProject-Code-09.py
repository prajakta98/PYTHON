from openpyxl import load_workbook
from openpyxl.utils import get_column_letter,column_index_from_string
wb=load_workbook('VITPython18-FinalProject-Workbook.xlsx')
sheet=wb["UserDetails"]
print(sheet.title)
username=input("Enter Full Name:")
list2=[]
def createlist(j):
    for i in range(2,7):
        list1=sheet.cell(row=i,column=j).value
        list2.append(list1)
    #print(list2)
    return list2
def validity(username,j):
    list2=createlist(j)
    if username in list2:
        return 1
    else:
        return 0
for k in range(0,3):
    n=validity(username,4)
    if n==1:
        break
    else:
        if k<2:
            print("Invalid Input")
            username=input("Enter Full Name:")
        else:
            print("program quit in stage 1")
            print(username.row)
            break
accountno=input("enter account number")
for m in range(0,3):
    n=validity(accountno,3)
    if n==1:
        break
    else:
        if m<2:
            print("Invalid Input")
            accountno=input("Enter account number:")
        else:
            print("program quit in stage 1")
            print(accountno.row)
            break

        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
  
           
    
       
