from openpyxl import load_workbook
wb=load_workbook('workbook-ex-1.xlsx')
wb.create_sheet(index=0,title="ex-sheet-1")
print(wb.sheetnames)
sheet=wb["ex-sheet-1"]
sheet['A1']="A"
print(sheet['A1'].value)
sheet['B1']="B"
print(sheet['B1'].value)
sheet['C1']="SUM"
print(sheet['C1'].value)
a=1
for i in range(2,8):
    for l in range(1,3):
        sheet.cell(row=i,column=l).value=a
        print(a)
        a=a+1
for s in range(2,8)
    k=
    sheet.cell(row=i,column=3).value=